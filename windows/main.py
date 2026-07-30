import csv, json, os, queue, sqlite3, threading, time, shutil, zipfile
from datetime import datetime, timedelta
from pathlib import Path
from urllib import request
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from openpyxl import load_workbook, Workbook

APP = "GSMS SMS v1.0"
DB = Path(os.getenv("APPDATA", Path.home())) / "GSMS_SMS_v1" / "gsms.db"
DB.parent.mkdir(parents=True, exist_ok=True)
DEFAULT_PASSWORD = "1234"
MAX_15, MAX_HOUR, MAX_DAY = 150, 250, 750

class DBStore:
    def __init__(self):
        self.db=sqlite3.connect(DB,check_same_thread=False); self.lock=threading.Lock()
        self.db.execute("CREATE TABLE IF NOT EXISTS settings(k TEXT PRIMARY KEY,v TEXT NOT NULL)")
        self.db.execute("CREATE TABLE IF NOT EXISTS contacts(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT,phone TEXT,group_name TEXT,kind TEXT)")
        self.db.execute("CREATE TABLE IF NOT EXISTS staff(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT,phone TEXT,designation TEXT)")
        self.db.execute("CREATE TABLE IF NOT EXISTS sent_log(id INTEGER PRIMARY KEY AUTOINCREMENT,phone TEXT,message TEXT,created_at TEXT,status TEXT)")
        self.db.execute("CREATE TABLE IF NOT EXISTS templates(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT UNIQUE,message TEXT)")
        self.db.commit()
        if self.get("password") is None:self.set("password",DEFAULT_PASSWORD)
        if self.get("android_url") is None:self.set("android_url","http://192.168.1.100:8765")
        if self.get("token") is None:self.set("token","")
    def get(self,k):
        row=self.db.execute("SELECT v FROM settings WHERE k=?",(k,)).fetchone(); return row[0] if row else None
    def set(self,k,v):
        with self.lock:self.db.execute("INSERT INTO settings(k,v) VALUES(?,?) ON CONFLICT(k) DO UPDATE SET v=excluded.v",(k,v)); self.db.commit()
    def contacts(self):return self.db.execute("SELECT id,name,phone,group_name,kind FROM contacts ORDER BY name").fetchall()
    def add_contact(self,n,p,g,k):
        with self.lock:self.db.execute("INSERT INTO contacts(name,phone,group_name,kind) VALUES(?,?,?,?)",(n,p,g,k));self.db.commit()
    def update_contact(self,i,n,p,g,k):
        with self.lock:self.db.execute("UPDATE contacts SET name=?,phone=?,group_name=?,kind=? WHERE id=?",(n,p,g,k,i));self.db.commit()
    def delete_contacts(self,ids):
        with self.lock:
            self.db.executemany("DELETE FROM contacts WHERE id=?",[(i,) for i in ids]);self.db.commit()
    def staff(self):return self.db.execute("SELECT id,name,phone,designation FROM staff ORDER BY name").fetchall()
    def add_staff(self,n,p,d):
        with self.lock:self.db.execute("INSERT INTO staff(name,phone,designation) VALUES(?,?,?)",(n,p,d));self.db.commit()
    def update_staff(self,i,n,p,d):
        with self.lock:self.db.execute("UPDATE staff SET name=?,phone=?,designation=? WHERE id=?",(n,p,d,i));self.db.commit()
    def delete_staff(self,ids):
        with self.lock:self.db.executemany("DELETE FROM staff WHERE id=? ",[(i,) for i in ids]);self.db.commit()
    def logs(self):return self.db.execute("SELECT id,phone,message,created_at,status FROM sent_log ORDER BY id DESC").fetchall()
    def log(self,p,m,s):
        with self.lock:self.db.execute("INSERT INTO sent_log(phone,message,created_at,status) VALUES(?,?,?,?)",(p,m,datetime.now().isoformat(timespec="seconds"),s));self.db.commit()
    def templates(self):return self.db.execute("SELECT id,name,message FROM templates ORDER BY name").fetchall()
    def add_template(self,n,m):
        with self.lock:self.db.execute("INSERT OR REPLACE INTO templates(name,message) VALUES(?,?)",(n,m));self.db.commit()
    def delete_template(self,i):
        with self.lock:self.db.execute("DELETE FROM templates WHERE id=?",(i,));self.db.commit()
    def backup(self,target):
        with self.lock:self.db.commit(); shutil.copy2(DB,target)
    def restore(self,source):
        with self.lock:
            self.db.close(); shutil.copy2(source,DB); self.db=sqlite3.connect(DB,check_same_thread=False); self.db.execute("PRAGMA foreign_keys=ON")

class RateLimiter:
    def __init__(self):self.times=[];self.lock=threading.Lock()
    def can_send(self):
        now=time.time()
        with self.lock:
            self.times=[t for t in self.times if now-t<86400]
            return sum(now-t<900 for t in self.times)<MAX_15 and sum(now-t<3600 for t in self.times)<MAX_HOUR and len(self.times)<MAX_DAY
    def wait_seconds(self):
        now=time.time()
        with self.lock:
            self.times=[t for t in self.times if now-t<86400]; waits=[]
            if sum(now-t<900 for t in self.times)>=MAX_15:waits.append(900-(now-min(t for t in self.times if now-t<900)))
            if sum(now-t<3600 for t in self.times)>=MAX_HOUR:waits.append(3600-(now-min(t for t in self.times if now-t<3600)))
            if len(self.times)>=MAX_DAY:waits.append(86400-(now-min(self.times)))
            return max(waits) if waits else 0
    def mark(self):
        with self.lock:self.times.append(time.time())

class App:
    def __init__(self,root):
        self.root=root;self.store=DBStore();self.limiter=RateLimiter();self.pause=threading.Event();self.pause.set();self.q=queue.Queue()
        root.title(APP+" — GHSS Utror Swat");root.geometry("1180x760");self.build_login()
    def clear(self):
        for w in self.root.winfo_children():w.destroy()
    def build_login(self):
        self.clear();f=ttk.Frame(self.root,padding=45);f.pack(expand=True)
        ttk.Label(f,text="GSMS SMS v1.0",font=("Segoe UI",26,"bold")).pack(pady=8)
        ttk.Label(f,text="Govt. Higher Secondary School Utror Swat",font=("Segoe UI",14)).pack()
        ttk.Label(f,text="Principal Login",font=("Segoe UI",12)).pack(pady=(28,6));pw=ttk.Entry(f,show="*",width=28);pw.pack();pw.focus()
        ttk.Button(f,text="Login",command=lambda:self.login(pw.get())).pack(pady=15)
        ttk.Label(f,text="Default first-login password: 1234 (change it immediately)",foreground="#7a1f1f").pack()
    def login(self,p):
        if p==self.store.get("password"):self.build_main()
        else:messagebox.showerror(APP,"Incorrect password.")
    def build_main(self):
        self.clear();top=ttk.Frame(self.root,padding=10);top.pack(fill="x")
        ttk.Label(top,text="GSMS SMS v1.0",font=("Segoe UI",18,"bold")).pack(side="left")
        ttk.Label(top,text="  •  GHSS Utror Swat",font=("Segoe UI",11)).pack(side="left")
        for text,cmd in [("Backup",self.backup),("Restore",self.restore),("Settings",self.settings),("Change Password",self.change_password)]:
            ttk.Button(top,text=text,command=cmd).pack(side="right",padx=3)
        self.nb=ttk.Notebook(self.root);self.nb.pack(fill="both",expand=True,padx=10,pady=5)
        self.contacts_tab();self.staff_tab();self.templates_tab();self.history_tab();self.reports_tab();self.queue_tab()
        threading.Thread(target=self.worker,daemon=True).start()
    def contacts_tab(self):
        tab=ttk.Frame(self.nb,padding=8);self.nb.add(tab,text="Students / Parents")
        bar=ttk.Frame(tab);bar.pack(fill="x")
        ttk.Button(bar,text="Import Excel/CSV",command=self.import_contacts).pack(side="left")
        ttk.Button(bar,text="Add",command=lambda:self.contact_editor()).pack(side="left",padx=4)
        ttk.Button(bar,text="Edit",command=self.edit_contact).pack(side="left")
        ttk.Button(bar,text="Delete",command=self.delete_contact).pack(side="left",padx=4)
        ttk.Label(bar,text="Group:").pack(side="left",padx=(20,3));self.group_var=tk.StringVar(value="All");self.group_box=ttk.Combobox(bar,textvariable=self.group_var,state="readonly",width=18);self.group_box.pack(side="left");self.group_box.bind("<<ComboboxSelected>>",lambda e:self.refresh_contacts())
        self.ctree=ttk.Treeview(tab,columns=("id","name","phone","group","kind"),show="headings",selectmode="extended")
        for c,t,w in [("id","ID",45),("name","Name",190),("phone","Phone",140),("group","Class/Group",130),("kind","Type",110)]:self.ctree.heading(c,text=t);self.ctree.column(c,width=w)
        self.ctree.pack(fill="both",expand=True,pady=7);self.refresh_contacts()
        b=ttk.Frame(tab);b.pack(fill="x");ttk.Button(b,text="Select All",command=lambda:self.ctree.selection_set(self.ctree.get_children())).pack(side="left");ttk.Button(b,text="Compose SMS to Selected",command=lambda:self.compose_from_tree(self.ctree)).pack(side="right")
    def refresh_contacts(self):
        if not hasattr(self,"ctree"):return
        for x in self.ctree.get_children():self.ctree.delete(x)
        rows=self.store.contacts();groups=sorted({r[3] for r in rows if r[3]})
        self.group_box["values"]=["All"]+groups
        g=self.group_var.get()
        for r in rows:
            if g=="All" or r[3]==g:self.ctree.insert("","end",values=r)
    def contact_editor(self,row=None):
        w=tk.Toplevel(self.root);w.title("Contact");f=ttk.Frame(w,padding=15);f.pack()
        vals=row or ("","","","")
        labels=["Name","Phone","Class/Group","Type"];es=[]
        for i,l in enumerate(labels):
            ttk.Label(f,text=l).grid(row=i,column=0,sticky="w",pady=4);e=ttk.Entry(f,width=36);e.grid(row=i,column=1,pady=4);e.insert(0,str(vals[i+1]) if row else "");es.append(e)
        def save():
            v=[e.get().strip() for e in es]
            if not v[1]:return messagebox.showwarning(APP,"Phone number is required.",parent=w)
            if row:self.store.update_contact(row[0],*v)
            else:self.store.add_contact(*v)
            w.destroy();self.refresh_contacts()
        ttk.Button(f,text="Save",command=save).grid(row=4,column=1,sticky="e",pady=8)
    def edit_contact(self):
        s=self.ctree.selection()
        if len(s)!=1:return messagebox.showwarning(APP,"Select exactly one contact.")
        self.contact_editor(self.ctree.item(s[0],"values"))
    def delete_contact(self):
        s=self.ctree.selection()
        if not s:return
        if messagebox.askyesno(APP,f"Delete {len(s)} selected contact(s)?"):
            self.store.delete_contacts([int(self.ctree.item(i,"values")[0]) for i in s]);self.refresh_contacts()
    def compose_from_tree(self,tree):
        s=tree.selection()
        if not s:return messagebox.showwarning(APP,"Select at least one recipient.")
        w=tk.Toplevel(self.root);w.title("Compose SMS");w.geometry("600x420");f=ttk.Frame(w,padding=12);f.pack(fill="both",expand=True)
        ttk.Label(f,text=f"{len(s)} recipient(s) selected").pack(anchor="w")
        msg=tk.Text(f,height=9);msg.pack(fill="x",pady=8)
        ttk.Label(f,text="Template:").pack(anchor="w");tv=tk.StringVar();cb=ttk.Combobox(f,textvariable=tv,state="readonly",width=50);cb["values"]=[t[1] for t in self.store.templates()];cb.pack(anchor="w")
        def load(): 
            name=tv.get()
            for t in self.store.templates():
                if t[1]==name:msg.delete("1.0","end");msg.insert("1.0",t[2])
        ttk.Button(f,text="Load Template",command=load).pack(anchor="w",pady=3)
        def qsend():
            text=msg.get("1.0","end").strip()
            if not text:return
            for i in s:
                r=tree.item(i,"values");self.q.put((r[2],text,r[1]));self.queue_view.insert("end",f"Queued: {r[1]} — {r[2]}")
            self.status.set(f"Queued {len(s)} SMS(s).");w.destroy();self.nb.select(self.queue_tab_ref)
        ttk.Button(f,text="Queue SMS",command=qsend).pack(anchor="e",pady=10)
    def staff_tab(self):
        tab=ttk.Frame(self.nb,padding=8);self.nb.add(tab,text="Staff SMS")
        bar=ttk.Frame(tab);bar.pack(fill="x");ttk.Button(bar,text="Add Staff",command=self.add_staff).pack(side="left");ttk.Button(bar,text="Edit",command=self.edit_staff).pack(side="left",padx=4);ttk.Button(bar,text="Delete",command=self.delete_staff).pack(side="left")
        self.stree=ttk.Treeview(tab,columns=("id","name","phone","designation"),show="headings",selectmode="extended")
        for c,t,w in [("id","ID",45),("name","Name",220),("phone","Phone",160),("designation","Designation",220)]:self.stree.heading(c,text=t);self.stree.column(c,width=w)
        self.stree.pack(fill="both",expand=True,pady=7);self.refresh_staff();ttk.Button(tab,text="Compose SMS to Selected Staff",command=lambda:self.compose_from_tree(self.stree)).pack(anchor="e")
    def refresh_staff(self):
        if not hasattr(self,"stree"):return
        for x in self.stree.get_children():self.stree.delete(x)
        for r in self.store.staff():self.stree.insert("","end",values=r)
    def staff_editor(self,row=None):
        w=tk.Toplevel(self.root);w.title("Staff Member");f=ttk.Frame(w,padding=15);f.pack();vals=row or ("","","")
        es=[]
        for i,l in enumerate(["Name","Phone","Designation"]):
            ttk.Label(f,text=l).grid(row=i,column=0,sticky="w",pady=4);e=ttk.Entry(f,width=36);e.grid(row=i,column=1);e.insert(0,str(vals[i+1]) if row else "");es.append(e)
        def save():
            v=[e.get().strip() for e in es]
            if not v[1]:return
            if row:self.store.update_staff(row[0],*v)
            else:self.store.add_staff(*v)
            w.destroy();self.refresh_staff()
        ttk.Button(f,text="Save",command=save).grid(row=3,column=1,sticky="e",pady=8)
    def add_staff(self):self.staff_editor()
    def edit_staff(self):
        s=self.stree.selection()
        if len(s)==1:self.staff_editor(self.stree.item(s[0],"values"))
    def delete_staff(self):
        s=self.stree.selection()
        if s and messagebox.askyesno(APP,"Delete selected staff?"):self.store.delete_staff([int(self.stree.item(i,"values")[0]) for i in s]);self.refresh_staff()
    def templates_tab(self):
        tab=tk.Frame(self.nb);self.nb.add(tab,text="SMS Templates");f=ttk.Frame(tab,padding=10);f.pack(fill="both",expand=True)
        self.tlist=tk.Listbox(f);self.tlist.pack(side="left",fill="both",expand=True);self.tmsg=tk.Text(f,width=60,height=12);self.tmsg.pack(side="right",fill="both",expand=True,padx=10)
        b=ttk.Frame(f);b.pack(side="bottom",fill="x",pady=5);ttk.Button(b,text="New/Save Template",command=self.save_template).pack(side="left");ttk.Button(b,text="Delete",command=self.delete_template).pack(side="left",padx=4);self.refresh_templates()
        self.tlist.bind("<<ListboxSelect>>",self.load_template)
    def refresh_templates(self):
        if hasattr(self,"tlist"):self.tlist.delete(0,"end");[self.tlist.insert("end",t[1]) for t in self.store.templates()]
    def load_template(self,e=None):
        s=self.tlist.curselection()
        if not s:return
        t=self.store.templates()[s[0]];self.tmsg.delete("1.0","end");self.tmsg.insert("1.0",t[2])
    def save_template(self):
        name=simpledialog.askstring(APP,"Template name:")
        msg=self.tmsg.get("1.0","end").strip()
        if name and msg:self.store.add_template(name,msg);self.refresh_templates()
    def delete_template(self):
        s=self.tlist.curselection()
        if s:
            self.store.delete_template(self.store.templates()[s[0]][0]);self.refresh_templates();self.tmsg.delete("1.0","end")
    def history_tab(self):
        tab=ttk.Frame(self.nb,padding=8);self.nb.add(tab,text="SMS History");bar=ttk.Frame(tab);bar.pack(fill="x");ttk.Button(bar,text="Refresh",command=self.refresh_history).pack(side="left");ttk.Button(bar,text="Export CSV",command=self.export_history).pack(side="left",padx=4)
        self.htree=ttk.Treeview(tab,columns=("id","phone","message","time","status"),show="headings"); 
        for c,t,w in [("id","ID",45),("phone","Phone",140),("message","Message",430),("time","Date/Time",160),("status","Status",100)]:self.htree.heading(c,text=t);self.htree.column(c,width=w)
        self.htree.pack(fill="both",expand=True,pady=6);self.refresh_history()
    def refresh_history(self):
        if not hasattr(self,"htree"):return
        for x in self.htree.get_children():self.htree.delete(x)
        for r in self.store.logs():self.htree.insert("","end",values=r)
    def export_history(self):
        path=filedialog.asksaveasfilename(defaultextension=".csv",filetypes=[("CSV","*.csv")])
        if not path:return
        with open(path,"w",newline="",encoding="utf-8-sig") as f:
            wr=csv.writer(f);wr.writerow(["ID","Phone","Message","Date/Time","Status"]);wr.writerows(self.store.logs())
    def reports_tab(self):
        tab=ttk.Frame(self.nb,padding=20);self.nb.add(tab,text="Reports")
        self.report_var=tk.StringVar();ttk.Label(tab,textvariable=self.report_var,font=("Segoe UI",14)).pack(anchor="w",pady=10);ttk.Button(tab,text="Refresh Report",command=self.refresh_report).pack(anchor="w");self.refresh_report()
    def refresh_report(self):
        logs=self.store.logs();total=len(logs);ok=sum(1 for r in logs if r[4]=="accepted");fail=total-ok
        self.report_var.set(f"SMS activity\n\nTotal queued/sent attempts recorded: {total}\nAccepted by Android gateway: {ok}\nFailed gateway requests: {fail}\n\nSafety policy: {MAX_15}/15 min • {MAX_HOUR}/hour • {MAX_DAY}/24 hours")
    def queue_tab(self):
        self.queue_tab_ref=ttk.Frame(self.nb,padding=8);self.nb.add(self.queue_tab_ref,text="SMS Queue")
        ctl=ttk.Frame(self.queue_tab_ref);ctl.pack(fill="x");ttk.Button(ctl,text="Pause",command=self.pause.clear).pack(side="left");ttk.Button(ctl,text="Resume",command=self.pause.set).pack(side="left",padx=5);ttk.Button(ctl,text="Clear Queue",command=self.clear_queue).pack(side="left")
        self.status=tk.StringVar(value="Ready");ttk.Label(self.queue_tab_ref,textvariable=self.status).pack(anchor="w",pady=8);self.queue_view=tk.Listbox(self.queue_tab_ref,height=20);self.queue_view.pack(fill="both",expand=True);ttk.Label(self.queue_tab_ref,text="Controlled sending: 150/15 min • 250/hour • 750/24 hours").pack(anchor="w",pady=5)
    def clear_queue(self):
        while not self.q.empty():
            try:self.q.get_nowait();self.q.task_done()
            except queue.Empty:break
        self.queue_view.delete(0,"end");self.status.set("Queue cleared.")
    def worker(self):
        while True:
            phone,msg,name=self.q.get()
            try:
                self.pause.wait()
                while not self.limiter.can_send():
                    wait=self.limiter.wait_seconds();self.status_set(f"Safety limit reached. Waiting about {int(wait//60)+1} minute(s).");time.sleep(min(max(wait,1),60));self.pause.wait()
                ok,detail=self.send_android(phone,msg);self.store.log(phone,msg,"accepted" if ok else "failed")
                if ok:self.limiter.mark();self.status_set(f"Accepted for sending: {name} ({phone})")
                else:self.status_set(f"Failed: {name} — {detail}")
            finally:self.q.task_done()
    def status_set(self,t):self.root.after(0,lambda:self.status.set(t))
    def send_android(self,phone,msg):
        base=self.store.get("android_url").rstrip("/");token=self.store.get("token") or "";data=json.dumps({"phone":phone,"message":msg}).encode()
        req=request.Request(base+"/send",data=data,method="POST",headers={"Content-Type":"application/json","X-GSMS-Token":token})
        try:
            with request.urlopen(req,timeout=12) as r:return r.status==200,r.read().decode(errors="replace")
        except Exception as e:return False,str(e)
    def settings(self):
        w=tk.Toplevel(self.root);w.title("Gateway Settings");f=ttk.Frame(w,padding=20);f.pack();u=ttk.Entry(f,width=48);u.insert(0,self.store.get("android_url"));u.grid(row=0,column=1);ttk.Label(f,text="Android URL").grid(row=0,column=0)
        t=ttk.Entry(f,width=48);t.insert(0,self.store.get("token"));t.grid(row=1,column=1);ttk.Label(f,text="Pairing token").grid(row=1,column=0)
        ttk.Button(f,text="Save",command=lambda:(self.store.set("android_url",u.get().strip()),self.store.set("token",t.get().strip()),w.destroy())).grid(row=2,column=1,sticky="e",pady=10)
    def change_password(self):
        old=simpledialog.askstring(APP,"Current password",show="*")
        if old!=self.store.get("password"):return messagebox.showerror(APP,"Current password is incorrect.")
        new=simpledialog.askstring(APP,"New password",show="*")
        if new and len(new)>=4:self.store.set("password",new);messagebox.showinfo(APP,"Password changed.")
    def backup(self):
        p=filedialog.asksaveasfilename(defaultextension=".db",filetypes=[("GSMS backup","*.db")])
        if p:self.store.backup(p);messagebox.showinfo(APP,"Backup created successfully.")
    def restore(self):
        p=filedialog.askopenfilename(filetypes=[("GSMS backup","*.db")])
        if p and messagebox.askyesno(APP,"Restore this backup? Current data will be replaced."):
            self.store.restore(p);self.refresh_contacts();self.refresh_staff();self.refresh_history();self.refresh_templates();messagebox.showinfo(APP,"Backup restored.")
    def import_contacts(self):
        p=filedialog.askopenfilename(filetypes=[("Excel/CSV","*.xlsx *.xlsm *.csv"),("All files","*.*")])
        if not p:return
        try:
            if p.lower().endswith(".csv"):
                with open(p,newline="",encoding="utf-8-sig") as f:rows=list(csv.reader(f))
            else:rows=[list(r) for r in load_workbook(p,read_only=True,data_only=True).active.iter_rows(values_only=True)]
            if not rows:raise ValueError("No rows found")
            h=[str(x or "").strip().lower() for x in rows[0]]
            def idx(ns,d):
                for n in ns:
                    if n in h:return h.index(n)
                return d
            ni,pi,gi,ki=idx(["name","student name","student","parent name"],0),idx(["phone","mobile","mobile no","phone number","parent phone"],1),idx(["group","class","section"],2),idx(["type","kind","category"],3);count=0
            for r in rows[1:]:
                if len(r)<=max(ni,pi):continue
                name=str(r[ni] or "").strip();phone=str(r[pi] or "").strip()
                if phone:self.store.add_contact(name,phone,str(r[gi] or "") if gi<len(r) else "",str(r[ki] or "Student/Parent") if ki<len(r) else "Student/Parent");count+=1
            self.refresh_contacts();messagebox.showinfo(APP,f"Imported {count} contact(s).")
        except Exception as e:messagebox.showerror(APP,f"Import failed: {e}")
    def edit_staff(self):
        s=self.stree.selection()
        if len(s)==1:self.staff_editor(self.stree.item(s[0],"values"))
    def staff_editor(self,row=None):
        w=tk.Toplevel(self.root);w.title("Staff Member");f=ttk.Frame(w,padding=15);f.pack();vals=row or ("","","");es=[]
        for i,l in enumerate(["Name","Phone","Designation"]):
            ttk.Label(f,text=l).grid(row=i,column=0,sticky="w",pady=4);e=ttk.Entry(f,width=36);e.grid(row=i,column=1);e.insert(0,str(vals[i+1]) if row else "");es.append(e)
        def save():
            v=[e.get().strip() for e in es]
            if not v[1]:return messagebox.showwarning(APP,"Phone is required.",parent=w)
            (self.store.update_staff(row[0],*v) if row else self.store.add_staff(*v));w.destroy();self.refresh_staff()
        ttk.Button(f,text="Save",command=save).grid(row=3,column=1,sticky="e",pady=8)
    def add_staff(self):self.staff_editor()
    def delete_staff(self):
        s=self.stree.selection()
        if s and messagebox.askyesno(APP,"Delete selected staff?"):self.store.delete_staff([int(self.stree.item(i,"values")[0]) for i in s]);self.refresh_staff()
    def refresh_staff(self):
        if not hasattr(self,"stree"):return
        for x in self.stree.get_children():self.stree.delete(x)
        for r in self.store.staff():self.stree.insert("","end",values=r)
    def refresh_templates(self):
        if hasattr(self,"tlist"):
            self.tlist.delete(0,"end")
            for t in self.store.templates():self.tlist.insert("end",t[1])
if __name__=="__main__":
    root=tk.Tk();App(root);root.mainloop()
